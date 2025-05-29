from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django.utils import timezone
from .models import Category, Product
from .serializers import CategorySerializer, ProductSerializer, ProductCreateSerializer
from rest_framework.permissions import AllowAny
from io import TextIOWrapper
from inventory.models import Product, Category
from transactions.models import TransactionHistory
from datetime import timedelta
from rest_framework.views import APIView
from django.db.models import Sum, Count, Q, F
from django.db.models import Q, Case, When, Value, IntegerField
from .pagination import CustomPagination
import csv
import openpyxl
from .models import Product, User, Category
from django.utils.dateparse import parse_date
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction
from transactions.models import TransactionHistory, TransactionItem
from customers.models import Customer
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [AllowAny]

class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CustomPagination
    
    def get_queryset(self):
        user = self.request.user
        queryset = Product.objects.filter(owner=user)
        
        search = self.request.query_params.get('search')
        brands = self.request.query_params.getlist('brand')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        condition = self.request.query_params.get('condition')
        buyer = self.request.query_params.get('buyer')
        seller = self.request.query_params.get('seller')
        sort_by = self.request.query_params.get('sort_by', 'created_at')  # Default sort by created_at
        sort_direction = self.request.query_params.get('sort_direction', 'desc')  # Default descending
        is_transaction = self.request.query_params.get('is_transaction')
        
        if is_transaction:
            queryset = queryset.filter(quantity__gt=0)
        
        # Global search functionality
        if search:
            search_terms = search.strip().split()
            if search_terms:
                search_filters = []
                for term in search_terms:
                    term_filter = (
                        Q(model_name__icontains=term) | 
                        Q(product_id__icontains=term) | 
                        Q(category__name__icontains=term)
                    )
                    search_filters.append(term_filter)
                
                if search_filters:
                    combined_search_filter = search_filters[0]
                    for filter_item in search_filters[1:]:
                        combined_search_filter &= filter_item
                    queryset = queryset.filter(combined_search_filter)
        
        if brands:
            brand_queries = []
            for brand in brands:
                brand_words = brand.strip().split()
                if brand_words:
                    product_filters = []
                    for word in brand_words:
                        word_filter = Q(model_name__icontains=word) | Q(category__name__icontains=word)
                        product_filters.append(word_filter)
                    
                    if product_filters:
                        combined_filter = product_filters[0]
                        for filter_item in product_filters[1:]:
                            combined_filter &= filter_item
                        brand_queries.append(combined_filter)
            
            if brand_queries:
                brand_filter = brand_queries[0]
                for query in brand_queries[1:]:
                    brand_filter |= query
                queryset = queryset.filter(brand_filter)
        
        if start_date:
            queryset = queryset.filter(date_purchased__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(date_purchased__lte=end_date)
        
        if condition:
            queryset = queryset.filter(condition=condition)
        
        if buyer:
            queryset = queryset.filter(sold_source__icontains=buyer)
        
        if seller:
            queryset = queryset.filter(purchased_from__icontains=seller)
        
        valid_sort_fields = ['id', 'created_at']
        if sort_by in valid_sort_fields:
            if sort_direction.lower() == 'desc':
                queryset = queryset.annotate(
                    zero_quantity=Case(
                        When(quantity=0, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('zero_quantity', f'-{sort_by}')
            else:
                queryset = queryset.annotate(
                    zero_quantity=Case(
                        When(quantity=0, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('zero_quantity', sort_by)
        else:
            queryset = queryset.annotate(
                zero_quantity=Case(
                    When(quantity=0, then=Value(1)),
                    default=Value(0),
                    output_field=IntegerField()
                )
            ).order_by('zero_quantity', '-created_at')
        
        return queryset

    def get_serializer_class(self):
        if self.action == 'create':
            return ProductCreateSerializer
        return ProductSerializer
    
    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def create_product(self, request):
        serializer = ProductCreateSerializer(data=request.data, files=request.FILES)
        if serializer.is_valid():
            product = serializer.save(owner=request.user)
            response_serializer = ProductSerializer(product)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def unsold(self, request):
        unsold_products = self.get_queryset().filter(date_sold__isnull=True)
        serializer = self.get_serializer(unsold_products, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def sold(self, request):
        sold_products = self.get_queryset().filter(date_sold__isnull=False)
        serializer = self.get_serializer(sold_products, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['patch'])
    def mark_as_sold(self, request, pk=None):
        product = self.get_object()
        sold_data = {
            'date_sold': request.data.get('date_sold', timezone.now().date()),
            'sold_price': request.data.get('sold_price'),
            'sold_source': request.data.get('sold_source'),
            'status': 'sold'
        }
        
        serializer = self.get_serializer(product, data=sold_data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

    @action(detail=True, methods=['put', 'patch'])
    def update_product(self, request, pk=None):
        try:
            product = self.get_object()
            
            # Choose appropriate serializer based on update type
            if request.method == 'PUT':
                serializer = ProductSerializer(product, data=request.data)
            else:
                serializer = ProductSerializer(product, data=request.data, partial=True)
            
            if serializer.is_valid():
                if 'buying_price' in request.data and float(request.data['buying_price']) <= 0:
                    return Response(
                        {"buying_price": "Buying price must be greater than 0"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if request.data.get('status') == 'sold' and not request.data.get('sold_price'):
                    return Response(
                        {"sold_price": "Sold price is required when marking a product as sold"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                updated_product = serializer.save()
                response_serializer = ProductSerializer(updated_product)
                return Response(response_serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
        except Exception as e:
            return Response(
                {"error": f"Failed to update product: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['delete'])
    def delete_product(self, request, pk=None):
        try:
            product = self.get_object()
            model_name = product.model_name
            product.delete()
            
            return Response(
                {"message": f"Product '{model_name}' successfully deleted"},
                status=status.HTTP_200_OK
            )
        
        except Exception as e:
            return Response(
                {"error": f"Failed to delete product: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class DashboardStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        user = request.user
        today = timezone.now().date()
        seven_days_ago = today - timedelta(days=7)
        
        categories_count = Category.objects.filter(
            products__owner=user,
            products__created_at__gte=seven_days_ago
        ).distinct().count()

        total_products = Product.objects.filter(
            owner=user,
            created_at__gte=seven_days_ago
        ).count()
        
        sales_transactions = TransactionHistory.objects.filter(
            user=user,
            transaction_type='sale',
            date__gte=seven_days_ago
        )
        
        revenue = sum(transaction.total_sale_price for transaction in sales_transactions)
        
        transaction_items = TransactionItem.objects.filter(
            transaction__user=user,
            transaction__transaction_type='sale',
            transaction__date__gte=seven_days_ago
        )
        
        top_selling_count = transaction_items.count()
        top_selling_cost = sum(item.total_purchase_price for item in transaction_items)
        
        ordered_count = Product.objects.filter(
            owner=user,
            availability='reserved'
        ).count()
        
        not_in_stock = Product.objects.filter(
            owner=user,
            quantity=0,
            availability='in_stock'
        ).count()
        
        profit = sum(transaction.profit or 0 for transaction in sales_transactions)
        
        return Response({
            "categories": {
                "count": categories_count,
                "label": "Last 7 days"
            },
            "total_products": {
                "count": total_products,
                "label": "Last 7 days",
                "revenue": float(revenue)
            },
            "top_selling": {
                "count": top_selling_count,
                "label": "Last 7 days",
                "cost": float(top_selling_cost)
            },
            "low_stocks": {
                "ordered": ordered_count,
                "not_in_stock": not_in_stock
            },
            "profit": {
                "amount": float(profit),
                "label": "Last 7 days"
            }
        })
    

class ProductCSVUploadAPIView(APIView):

    @staticmethod
    def parse_decimal(value, default=0):
        if value is None:
            return Decimal(default)
        try:
            if isinstance(value, str):
                value = value.strip().replace(',', '')
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return Decimal(default)

    @staticmethod
    def to_date(value):
        if not value:
            return None
            
        if isinstance(value, datetime):
            return value.date()
            
        if not isinstance(value, str):
            return None
            
        value = value.strip()
        formats = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y']
        
        for date_format in formats:
            try:
                return datetime.strptime(value, date_format).date()
            except ValueError:
                continue
                
        return None
            
    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('excel_file')
        if not uploaded_file:
            return Response({'error': 'No file provided.'}, status=400)

        if uploaded_file.size == 0:
            return Response({'error': 'The uploaded file is empty (zero bytes).'}, status=400)

        filename = uploaded_file.name.lower()

        try:
            # Parse file based on extension
            if filename.endswith('.csv'):
                reader = csv.DictReader(TextIOWrapper(uploaded_file.file, encoding='utf-8'))
                rows = list(reader)
                if not rows:
                    return Response({'error': 'The uploaded CSV file is empty.'}, status=400)
                
                non_empty_rows = 0
                for row in rows:
                    has_data = False
                    for key, value in row.items():
                        if value and (not isinstance(value, str) or value.strip()):
                            has_data = True
                            break
                    if has_data:
                        non_empty_rows += 1
                
                if non_empty_rows == 0:
                    return Response({'error': 'The uploaded CSV file contains only empty rows.'}, status=400)
                
            elif filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb.active
                if sheet.max_row <= 1:
                    return Response({'error': 'The uploaded Excel file is empty or contains only headers.'}, status=400)
                    
                headers = [cell.value for cell in sheet[1] if cell.value]
                if not headers:
                    return Response({'error': 'The uploaded Excel file does not contain valid headers.'}, status=400)
                    
                rows = [
                    {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                    for row in sheet.iter_rows(min_row=2)
                ]
                if not rows:
                    return Response({'error': 'The uploaded Excel file does not contain any data rows.'}, status=400)
                
                non_empty_rows = 0
                for row in rows:
                    has_data = False
                    for key, value in row.items():
                        if value and (not isinstance(value, str) or value.strip()):
                            has_data = True
                            break
                    if has_data:
                        non_empty_rows += 1
                
                if non_empty_rows == 0:
                    return Response({'error': 'The uploaded Excel file contains only empty rows.'}, status=400)
            else:
                return Response({'error': 'Unsupported file format. Please upload CSV or Excel file.'}, status=400)
        except Exception as e:
            return Response({'error': f'Error processing file: {str(e)}'}, status=400)

        has_valid_data = False
        for row in rows:
            normalized_row = self._normalize_row(row)
            product_id = normalized_row.get('Reference')
            serial_number = normalized_row.get('Serial Number')
            model_name = normalized_row.get('Model Name')
            
            if product_id or serial_number or model_name:
                has_valid_data = True
                break
        
        if not has_valid_data:
            return Response({'error': 'The file does not contain any valid product data. Each product requires at least one of: Reference, Serial Number, or Model Name.'}, status=400)

        created = 0
        updated = 0
        errors = []
        user = self.request.user
        
        for index, row in enumerate(rows, start=2):
            try:
                with transaction.atomic():
                    normalized_row = self._normalize_row(row)
                    
                    product_id = normalized_row.get('Reference')
                    serial_number = normalized_row.get('Serial Number')
                    model_name = normalized_row.get('Model Name')
                    
                    if not (product_id or serial_number or model_name):
                        continue
                        
                    buy_price = self.parse_decimal(normalized_row.get('Buy Price'))
                    total_cost = self.parse_decimal(normalized_row.get('Total Cost'))
                    sell_price = self.parse_decimal(normalized_row.get('Sell Price'))
                    shipping_price = self.parse_decimal(normalized_row.get('Shipping'))
                    repair_cost = self.parse_decimal(normalized_row.get('Expense'))
                    
                    profit = sell_price - buy_price
                    profit_margin = int((profit / buy_price) * 100) if buy_price else 0
                    
                    brand_name = normalized_row.get('Brand') or 'Unnamed Product'
                    category_name = brand_name.strip()
                    category, _ = Category.objects.get_or_create(
                        name__iexact=category_name,
                        defaults={'name': category_name}
                    )
                    
                    date_purchased = self.to_date(normalized_row.get('Purchase Date'))
                    sold_date = self.to_date(normalized_row.get('Sold Date'))
                    
                    if not product_id:
                        product_id = f'custom-id-{index}'

                    existing_product = None
                    if product_id or serial_number:
                        q_filter = Q()
                        if product_id:
                            q_filter |= Q(product_id=product_id)
                        if serial_number:
                            q_filter |= Q(serial_number=serial_number)
                        
                        existing_products = Product.objects.filter(q_filter, owner=user)
                        if existing_products.exists():
                            existing_product = existing_products.first()

                    if not existing_product and product_id:
                        if Product.objects.filter(product_id=product_id).exists():
                            original_id = product_id
                            # product_id = f"{product_id}-{int(timezone.now().timestamp())}"
                            errors.append({
                                'row': index, 
                                'warning': f"Product ID '{original_id}' already exists. Created with modified ID '{product_id}'"
                            })

                    product_data = {
                        'owner': user,
                        'model_name': model_name or brand_name,
                        'product_id': product_id,
                        'serial_number': serial_number,
                        'category': category,
                        'profit_margin': profit_margin,
                        'availability': self.map_availability(normalized_row.get('Deal Status')),
                        'buying_price': buy_price,
                        'shipping_price': shipping_price,
                        'repair_cost': repair_cost,
                        'sold_price': sell_price,
                        'quantity': normalized_row.get('Quantity') or 1,
                        'date_purchased': date_purchased,
                        'date_sold': sold_date,
                        'source_of_sale': normalized_row.get('Sold To') or '',
                        'purchased_from': normalized_row.get('Bought From') or '',
                        'sold_source': normalized_row.get('Payment Sent account') or '',
                        'listed_on': normalized_row.get('Delivery Content') or '',
                        'wholesale_price': total_cost,
                    }

                    product = None
                    if existing_product:
                        for key, value in product_data.items():
                            setattr(existing_product, key, value)
                        existing_product.save()
                        product = existing_product
                        updated += 1
                    else:
                        product = Product.objects.create(**product_data)
                        created += 1
                    
                    # Create transaction if sell price and sold date are present
                    if buy_price and date_purchased:
                        # Check if a purchase transaction already exists for this product
                        if not TransactionItem.objects.filter(
                            product=product, 
                            transaction__transaction_type='purchase'
                        ).exists():
                            purchase_transaction_data = {
                                'user': user,
                                'name_of_trade': f"{product.model_name} Purchase",
                                'transaction_type': 'purchase',
                                'date': date_purchased,
                                'purchase_price': buy_price,
                                'expenses': {
                                    'shipping': float(shipping_price) if shipping_price else 0,
                                    'repair_cost': float(repair_cost) if repair_cost else 0
                                }
                            }
                            
                            # Get supplier if available
                            bought_from = normalized_row.get('Bought From')
                            if bought_from:
                                supplier, _ = Customer.objects.get_or_create(
                                    name=bought_from,
                                    defaults={
                                        'user': user,
                                        'is_supplier': True
                                    }
                                )
                                purchase_transaction_data['customer'] = supplier
                            
                            # Create purchase transaction
                            purchase_transaction = TransactionHistory.objects.create(
                                **purchase_transaction_data
                            )
                            
                            # Create purchase transaction item
                            TransactionItem.objects.create(
                                transaction=purchase_transaction,
                                product=product,
                                quantity=product.quantity or 1,
                                purchase_price=buy_price
                            )
                    
                    # Create sale transaction if sell price and sold date are present
                    if sell_price and sold_date:
                        # Check if a sale transaction already exists for this product
                        if not TransactionItem.objects.filter(
                            product=product, 
                            transaction__transaction_type='sale'
                        ).exists():
                            sale_transaction_data = {
                                'user': user,
                                'name_of_trade': f"{product.model_name} Sale",
                                'transaction_type': 'sale',
                                'date': sold_date,
                                'sale_price': sell_price,
                                'expenses': {
                                    'repair_cost': float(repair_cost) if repair_cost else 0
                                }
                            }
                            
                            # Get customer if available
                            sold_to = normalized_row.get('Sold To')
                            if sold_to:
                                customer, _ = Customer.objects.get_or_create(
                                    name=sold_to,
                                    defaults={'user': user}
                                )
                                sale_transaction_data['customer'] = customer
                            
                            # Set sale category if available
                            sale_category = normalized_row.get('Sale Category')
                            if sale_category and sale_category.lower() in [choice[0] for choice in TransactionHistory.SALE_CATEGORY_CHOICES]:
                                sale_transaction_data['sale_category'] = sale_category.lower()
                            
                            # Create sale transaction
                            sale_transaction = TransactionHistory.objects.create(
                                **sale_transaction_data
                            )
                            
                            # Create sale transaction item
                            TransactionItem.objects.create(
                                transaction=sale_transaction,
                                product=product,
                                quantity=1,
                                sale_price=sell_price,
                                purchase_price=buy_price  # Include original purchase price
                            )
                    
            except Exception as e:
                errors.append({'row': index, 'error': str(e)})

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'total_processed': created + updated,
            'total_rows': len(rows),
        }, status=201)

    def _normalize_row(self, row):
        return {
            key.strip() if isinstance(key, str) else key: 
            (value.strip() if isinstance(value, str) else value)
            for key, value in row.items()
            if key is not None
        }

    def map_availability(self, deal_status):
        if not deal_status:
            return 'in_stock'
            
        if isinstance(deal_status, str):
            deal_status = deal_status.strip().lower()
            
        mapping = {
            'sold': 'sold',
            'in stock': 'in_stock',
            'instock': 'in_stock',
            'reserved': 'reserved',
            'in repair': 'in_repair',
            'repair': 'in_repair'
        }
        return mapping.get(deal_status.lower(), 'in_stock')